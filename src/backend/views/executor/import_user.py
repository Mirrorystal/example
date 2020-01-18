from os import remove
from . import executor
from mongoengine.errors import NotUniqueError
from openpyxl import load_workbook
from models.setting import Setting
from models.task import Task
from models.user import User
from views.task import task_status_storage


def execute(task_id):
    return executor.submit(import_users, task_id)


def import_users(task_id):
    task = Task.objects(id=task_id).first()
    if task is None:
        return

    task.status = Task.STATUS_RUNNING
    task.save()

    wb = load_workbook(task.file.path)

    """计算需要导入用户的总数"""
    sheets = wb.sheetnames
    user_count = 0
    for sheet in sheets:
        ws = wb[sheet]
        user_count += ws.max_row - 1

    """缓存当前导入进度，用于前端调用接口获取"""
    storage_key = str(task_id)
    storage = task_status_storage[storage_key] = {
        'success': 0,
        'failed': 0,
        'total': user_count
    }

    settings = Setting.objects().first()
    if settings is None:
        settings = Setting()
        settings.save()

    failed_record = []
    for user, meta in load_users(wb, sheets):
        user.password = 'Abc123__'
        if user.department not in settings.departments:
            settings.update(add_to_set__departments=user.department)
            settings.reload()
        if user.position not in settings.positions:
            settings.update(add_to_set__positions=user.position)
            settings.reload()
        try:
            user.save()
            storage['success'] += 1
        except NotUniqueError:
            storage['failed'] += 1
            meta['msg'] = '该邮箱已被注册'
            failed_record.append(meta)
        except Exception as error:
            storage['failed'] += 1
            meta['msg'] = error.errors
            failed_record.append(meta)

    wb.close()

    task.status = Task.STATUS_COMPLETED
    task.result = {
        'success': storage['success'],
        'failed': storage['failed'],
        'failed_record': failed_record,
        'total': storage['total']
    }
    task.save()
    task_status_storage.pop(storage_key)
    remove(task.file.path)


def load_users(wb, sheets):

    for sheet in sheets:
        ws = wb[sheet]
        rows = ws.values
        header = next(rows)
        line_count = 1

        """遍历sheet中的user"""
        for row in rows:
            user = User()
            line_count += 1
            for i, h in enumerate(header):
                if h == '姓名':
                    user.username = str(row[i])
                elif h == '邮箱':
                    user.email = str(row[i])
                elif h == '职位':
                    user.position = str(row[i])
                elif h == '部门':
                    user.department = str(row[i])
                elif h == '电话号码':
                    user.phone = str(row[i])
            yield user, { 'line': line_count, 'sheet': sheet }


